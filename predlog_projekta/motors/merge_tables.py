import os
import csv
import openpyxl
from openpyxl import Workbook

TABLES_DIR = "tables"
OUTPUT_FILE = "all_tables.xlsx"

def read_csv(path):
    with open(path, newline="", encoding="utf-8") as f:
        return list(csv.reader(f))

def read_xlsx(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    result = []
    for sheet in wb.worksheets:
        merged = {}
        for merge_range in sheet.merged_cells.ranges:
            value = sheet.cell(merge_range.min_row, merge_range.min_col).value
            for row in range(merge_range.min_row, merge_range.max_row + 1):
                for col in range(merge_range.min_col, merge_range.max_col + 1):
                    merged[(row, col)] = value

        rows = []
        for row_idx, row in enumerate(sheet.iter_rows(values_only=False), 1):
            cells = []
            for col_idx, cell in enumerate(row, 1):
                if (row_idx, col_idx) in merged:
                    cells.append(merged[(row_idx, col_idx)])
                else:
                    cells.append(cell.value)
            if any(c is not None for c in cells):
                rows.append(cells)
        if rows:
            result.append(rows)
    return result

def sanitize_sheet_name(name):
    # Excel sheet names max 31 chars, no special chars
    for ch in r'\/*?:[]':
        name = name.replace(ch, "")
    return name[:31]

out_wb = Workbook()
out_wb.remove(out_wb.active)

files = sorted(os.listdir(TABLES_DIR))

for filename in files:
    path = os.path.join(TABLES_DIR, filename)
    stem = os.path.splitext(filename)[0]

    if filename.endswith(".csv"):
        rows = read_csv(path)
        if not rows:
            continue
        ws = out_wb.create_sheet(title=sanitize_sheet_name(stem))
        for row in rows:
            ws.append(row)

    elif filename.endswith(".xlsx"):
        tables = read_xlsx(path)
        if not tables:
            continue
        if len(tables) == 1:
            ws = out_wb.create_sheet(title=sanitize_sheet_name(stem))
            for row in tables[0]:
                ws.append(row)
        else:
            for idx, rows in enumerate(tables, 1):
                title = sanitize_sheet_name(f"{stem} t{idx}")
                ws = out_wb.create_sheet(title=title)
                for row in rows:
                    ws.append(row)

out_wb.save(OUTPUT_FILE)
print(f"Done -> {OUTPUT_FILE}")